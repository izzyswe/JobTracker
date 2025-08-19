from openpyxl import Workbook

wb = Workbook()
# get active worksheet
ws = wb.active

# all the neccesary columns
colName = ["id", "Company Name", "Position", "Address", "CV or Resume", "Web Link", "Status", "Date Applied", "Deadline"]
columns = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']
upperCol = [item.upper() for item in columns]


## PREVIOUS GENERATE TITLE
# [FIXED FUTURE PROBLEM] this has a fixed column with new data in every row
def generateTitles():
    for i in range(1, 11):
        cell = columns[0] + str(i)
        ws[cell] = "oof"

# CONCEPT CODE
# for i in rows:
#     if i == 0 :
#         continue
#     for j in columns:
#         print(j, i)

# x = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o']
# y = [1, 2, 3, 4, 5, 6, 7, 8, 9, 0]
# spreadsheet = [
#     ['a','b','c','d','e','f','g','h','i','j','k','l','m'], 
#     [1, 2, 3, 4, 5, 6, 7, 8, 9, 0]
# ]
# output b2
# print(spreadsheet[0][1],  spreadsheet[1][1], sep='')
# for i in x:
#     print(i)
# for i in y:
#     print(i)
#     if(i == 9):
#         print()
# for j in spreadsheet[1]:
#     print(spreadsheet[0][0] + str(j))

# how to add item in a dictionary loop
# # key() for keys
# # values() for values
# # items() for both key and value pair
#
# my_dict = {}
# for i in range(5):
#     key = f"item{{{i}}}"
#     value = i * 10
#     my_dict[key] = value  # Adds or updates the item
# print(my_dict)

# documentation Code
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# # ws['A2'] = datetime.datetime.now()
