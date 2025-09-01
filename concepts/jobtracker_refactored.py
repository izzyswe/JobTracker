"""
    Automation: Job Tracker
    Github: Izzyswe
    Name: Isaac

    Date: August 18, 2025

"""
import os
import pyperclip as pc
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime

class JobTracker:
    def __init__(self, filename="Job-Tracker.xlsx"):
        self.filename = filename
        self.col_names = ["id", "Company Name", "Position", "Address", "CV or Resume", "Web Link", "Status", "Date Applied", "Deadline"]
        self.wb = self._check_worksheet()
        self.ws = self.wb.active

    def _check_worksheet(self):
        if os.path.exists(self.filename):
            wb = load_workbook(self.filename)
        else:
            wb = Workbook()
            ws = wb.active
            self._generate_titles(ws)
        return wb

    def _generate_titles(self, worksheet):
        heading_fill = PatternFill(fill_type='solid', start_color='87f5a4', end_color='87f5a4')
        heading_font = Font(size=16, bold=True)
        for i, n in enumerate(self.col_names, start=1):
            cell = get_column_letter(i) + "2"
            worksheet[cell] = n

        for col_idx, header in enumerate(self.col_names, start=1):
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = len(header) + 5

        worksheet.merge_cells('A1:I1')
        worksheet['A1'] = "Job Tracker"
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A1'].font = heading_font
        worksheet['A1'].fill = heading_fill

    def add_job_details(self):
        if os.name == 'nt':
            os.system('cls')
        else:
            os.system('clear')
        print("JOB ENTRY")
        row_data = []
        has_applied_status = has_not_applied_status = is_currently_applying_status = False
        for i in self.col_names:
            if i == self.col_names[5]:
                val_input = pc.paste()
                print(f"Enter {i}: \n> {val_input}")
            elif i == self.col_names[6]:
                val_input = input(f"Enter {i}: \n> ")
                if val_input.lower() in ("applied", "sent", "a"):
                    has_applied_status = True
                elif val_input.lower() in ("not applied", "n"):
                    has_not_applied_status = True
                elif val_input.lower() == "applying":
                    is_currently_applying_status = True
            elif i == self.col_names[7] or i == self.col_names[8]:
                while True:
                    val_input = input(f"Enter {i} (YYYYMMDD)\n> ")
                    try:
                        val_input = datetime.strptime(val_input, "%Y%m%d").date()
                        break
                    except ValueError:
                        print("Invalid Format. Use YYYYMMDD")
            else:
                val_input = input(f"Enter {i}\n> ")
            row_data.append(val_input)

        self.ws.append(row_data)

        if has_applied_status:
            status_clr = PatternFill(fill_type='solid', start_color='005ef5', end_color='005ef5')
            cell_fill = f"{get_column_letter(self.col_names.index('Status') + 1)}{self.ws.max_row}"
            self.ws[cell_fill].fill = status_clr
        elif has_not_applied_status:
            status_clr = PatternFill(fill_type='solid', start_color='f50039', end_color='f50039')
            cell_fill = f"{get_column_letter(self.col_names.index('Status') + 1)}{self.ws.max_row}"
            self.ws[cell_fill].fill = status_clr
        elif is_currently_applying_status:
            status_clr = PatternFill(fill_type='solid', start_color='26ff40', end_color='26ff40')
            cell_fill = f"{get_column_letter(self.col_names.index('Status') + 1)}{self.ws.max_row}"
            self.ws[cell_fill].fill = status_clr

        print("-- Entry Added! --\n")

    def search_job_detail(self):
        if os.name == 'nt':
            os.system('cls')
        else:
            os.system('clear')
        print("SEARCH JOB APPLICATION\n")
        
        search_options = {
            "1": "id",
            "2": "Company Name",
            "3": "Position",
            "4": "Status",
        }

        print("Search by the Following:")
        for key, value in search_options.items():
            print(f"{key}. {value}")
        
        search_input = input(">>>> ")

        if search_input in search_options:
            search_column = search_options[search_input]
            search_term = input(f"Enter the {search_column} to search for: ")
            
            results = []
            for row in self.ws.iter_rows(min_row=3, values_only=True):
                row_dict = dict(zip(self.col_names, row))
                if row_dict.get(search_column) and str(row_dict[search_column]).lower() == search_term.lower():
                    results.append(row_dict)
            
            if results:
                print("\n--- Search Results ---")
                for res in results:
                    for key, val in res.items():
                        print(f"{key}: {val}")
                    print("--------------------")
            else:
                print("No results found.")
        else:
            print("Invalid option.")


    def update_job_detail(self):
        if os.name == 'nt':
            os.system('cls')
        else:
            os.system('clear')
        print("UPDATE JOB DETAIL - Coming soon")
        pass

    def main_menu(self):
        while True:
            print("\n|  JOB TRACKER MENU  | \n\n" +
                  "1. Enter Job Entry \n" +
                  "2. Search Job Entry \n" +
                  "3. Update Job Entry \n" +
                  "4. Quit Program")
            opts = input("> ")
            print()
            if opts == "4":
                print("Exiting program.")
                break

            match opts:
                case "1":
                    self.add_job_details()
                case "2":
                    self.search_job_detail()
                case "3":
                    self.update_job_detail()
                case _:
                    print("Invalid Input")

    def save_workbook(self):
        self.wb.save(self.filename)

if __name__ == "__main__":
    try:
        tracker = JobTracker()
        tracker.main_menu()
    except KeyboardInterrupt:
        print("\n")
    finally:
        tracker.save_workbook()
        print("workload saved")
