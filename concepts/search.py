from openpyxl import load_workbook
from datetime import datetime
import re
# CONSTANTS
COLNAME = ["id", "Company Name", "Position", "Address", "CV or Resume", "Web Link", "Status", "Date Applied", "Deadline"]
COLUMNS = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']

WB = load_workbook("Job-Tracker.xlsx")
WS = WB.active


def searchByDate():
    isDateFormatted = True
    dateFormatStartInput = None
    dateFormatEndInput = None

    while isDateFormatted:
        getDateStartInput = input("date start: \n>")
        try:
            dateFormatStartInput = datetime.strptime(getDateStartInput, "%Y%m%d").date()
            break
        except ValueError:
            print("enter valid date (YYYYMMDD)")

    while isDateFormatted:
        getDateEndInput = input("date end: \n>")
        try:
            dateFormatEndInput = datetime.strptime(getDateEndInput, "%Y%m%d").date()
            break
        except ValueError:
            print("enter valid date (YYYYMMDD)")

    print(f"Searched: {dateFormatStartInput} - {dateFormatEndInput} \n")

    for row in WS.iter_rows(min_row=2, values_only=True):
        cellValue = row[7]
        if dateFormatStartInput <= dateFormatEndInput:
            print(f"Date ------- {cellValue}")
            print(f"Matches ----  {row[1], row[2], row[6]} \n")


def searchByID():
    isIDNumber = True

    getID: int = input("enter ID: \n>")
    while isIDNumber:
        if not getID.isnumeric():
            print("error: not a numerical value")
            getID: int = input("enter ID: \n")
            break
        else:
            print("WORKING")
            break

    print(f"Searched ID: {getID}")
    for row in WS.iter_rows(min_row=1, values_only=True):
        cellValue = row[0]
        if cellValue == getID:
            print(f"Data: {cellValue}")
            print(f"Matches: {row[1], row[2]}")


# def searchByLinks():
#
#     getLink = input("enter relevant link: \n> ")
#     website_URL = "/^(https?:\/\/(?:www\.)?|www\.)[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+(?:[\/\w .-]*)*\/?$/gm"
#     regSearch = re.search(website_URL, getLink)
#     print(regSearch)
#     pass


def searchByCompany():
    isCompanyStr = True
    getCompanyName = input("enter Company Name: \n> ")
    while isCompanyStr:
        if not getCompanyName.isalpha():
            print("Not a numerical value")
            getCompanyName = input("enter Company Name: \n> ")
            break
        else:
            print(f"Searched {getCompanyName}")
            break

        for row in WS.iter_rows(min_row=1, values_only=True):
            cellValue = row[1]
            if cellValue == getCompanyName:
                print(f"Data: {cellValue}")
                print(f"Matches: {row[1], row[2]}")
            else:
                print("Not Found")

def searchByFields(field_index, field_name):
    isFieldType = True
    search_term = input(f"Enter {field_name}:\n> ")
    found = False
    for row in WS.iter_rows(min_row=2, values_only=True):
        if row[field_index] == search_term:
            print(f"Match found: {row}")
            found = True
    if not found:
        print("No match found.")

# searchByCompany()
inputfield = input("pick a column"+
                   "1. Company")

match inputfield:
    case "1"
searchByFields()
