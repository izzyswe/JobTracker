'''
    Automation: Job Tracker
    Github: Izzyswe
    Name: Isaac

    Date: August 18, 2025
'''
import os
import sys
import pyperclip as pc
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
from jobSearch import jobSearch

# ____TODO____ [✔] - copy this to mark things as done

# [ fix, ONGOING ]
# 1. fix getJobDetails function
#   - refactor if statements, repetition == bad
#   - fix Pattern Fill

# [started, ONGOING]
# 2. work on Search JobDetail Function
#   - date range search ✔
#   - implemenet id, company, positon, status search
#   - implement power search

# [NOT STARTED, HOLD]
# 3. work on updateJobDetail

# [ FUTURE FEATURE WANTS ]
# 1. Refactor main to OOP
# 1. Add GUI (Tkinter or PyQT)
# 2. DRAG N DROP TXT FILE CONVERSION (tkinterdnd2)
# 3. WEB SCRAP LIVE CHANGES (BeautifulSoup2 or Selenium)


# all the neccesary columns
colName = ["id", "Company Name", "Position", "Address", "CV or Resume", "Web Link", "Status", "Date Applied", "Deadline"]
columns = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i']
upperCol = [item.upper() for item in columns]

# CONSTANT DICTIONARIES
APPLICATION_STATUS = {
    "hasAppliedStatus": PatternFill(fill_type='solid', start_color='005ef5', end_color='005ef5', fgColor='ffffff'),
    "hasNotAppliedStatus": PatternFill(fill_type='solid', start_color='f50039', end_color='f50039', fgColor='ffffff'),
    "isCurrentlyApplyingStatus": PatternFill(fill_type='solid', start_color='26ff40', end_color='26ff40')
}

STATUS_INPUT = {
    "applied": ("applied", "sent", "a"),
    "not applied": ("not applied", "n"),
    "applying": "applying"
}

# placeholders
#        cellFill = f"{get_column_letter(colName.index("Status") + 1)}{worksheet.max_row}",
# cellFill = f"{get_column_letter(colName.index("Status") + 1)}{worksheet.max_row}"   worksheet[cellFill].fill = status_clr,
#        cellFill = f"{get_column_letter(colName.index("Status") + 1)}{worksheet.max_row}"    worksheet[cellFill].fill = status_clr


def checkOs():
    os.system('cls') if os.name == 'nt' else os.system('clear')


# [ADDED] KEPT OVERRIDING SO THIS FUNCTION WILL CHECK IF IT EXISTS
# OR AUTOGENERATE A NEW SPREADSHEET IF IT DOES NOT EXIST TO PREVENT THE ORGINAL ISSUE
def checkWorksheet(filename="Job-Tracker.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        ws = wb.active
        generateTitles(ws)
    return wb


# [FIXED] CURRENT CODE IS WAY MORE SIMPLER THAN PREVIOUS CODE
def getJobDetails(worksheet):
    checkOs()
    print("JOB ENTRY")
    rowData = []
    # initialize the variable to solve UnboundLocalError, they were used before they were assigned
    hasAppliedStatus = hasNotAppliedStatus = isCurrentlyApplyingStatus = False
    # loop throughout the colName list to get input for every column
    for i in colName:
        ## [REFACTOR] TOO MANY FRIGGIN IF-STATEMENTS, BAD CODE
        if i == colName[5]:
            valInput = pc.paste()
            print(f"Enter {i}: \n> {valInput}")
        elif i == colName[6]:
            valInput = input(f"Enter {i}: \n> ").lower()
            # turned it to a tuple for concise reading and readonly purposes
            if valInput.lower() == ("applied", "sent", "a"):
                hasAppliedStatus = True
            elif valInput.lower() == ("not applied", "n"):
                hasNotAppliedStatus = True
            elif valInput.lower() == "applying":
                isCurrentlyApplyingStatus = True
        elif i == colName[7] or i == colName[8]:
            while True:
                valInput = input(f"Enter {i} (YYYYMMDD)\n> ")
                try:
                    valInput = datetime.strptime(valInput, "%Y%m%d").date()
                    break
                except ValueError:
                    print("Invalid Format. Use YYYYMMDD")
        else:
            valInput = input(f"Enter {i}\n> ").lower()
        # append every value into the rowData array
        rowData.append(valInput)

    # in the worksheet argument, append the rowData in openpyxl ws
    worksheet.append(rowData)

    ## [REFACTOR] THIS CAN BE MORE CONCISE, ITS REPETITIVE.
    # AFTER ALL THE ROW ENTRY HAS BEEN APPLIED
    if hasAppliedStatus:
        # [ FIXED ]: Use the correct string "Status" to find the column, not the old 'i' variable.
        # [ FIXED ]: Use worksheet.max_row, not max_row + 1.
        status_clr = PatternFill(fill_type='solid', start_color='005ef5', end_color='005ef5', fgColor='ffffff')
        cellFill = f"{get_column_letter(colName.index("Status") + 1)}{worksheet.max_row}"
        worksheet[cellFill].fill = status_clr
    elif hasNotAppliedStatus:
        status_clr = PatternFill(fill_type='solid', start_color='f50039', end_color='f50039', fgColor='ffffff')
        cellFill = f"{get_column_letter(colName.index("Status") + 1)}{worksheet.max_row}"
        worksheet[cellFill].fill = status_clr
    elif isCurrentlyApplyingStatus:
        status_clr = PatternFill(fill_type='solid', start_color='26ff40', end_color='26ff40')
        cellFill = f"{get_column_letter(colName.index("Status") + 1)}{worksheet.max_row}"
        worksheet[cellFill].fill = status_clr

    print("-- Entry Added! --\n")


def searchJobDetail(worksheet):
    checkOs()
    job_searcher = jobSearch(worksheet)  # Create instance with worksheet
    print("SEARCH JOB APPLICATION\n")
    # rowData = []
    searchInput = input("Search by the Following: \n" +
                        "bk to go back \n\n" +
                        "1. ID \n" +
                        "2. Company Name \n" +
                        "3. Position \n" +
                        "4. Status \n" +
                        "5. Date Range \n>>>> ")
    match searchInput:
        case "1":
            job_searcher.searchByID()
        case "2":
            job_searcher.searchByCompany()
        case "3":
            job_searcher.searchbyField(2, "Positon")
        case "4":
            pass
        case "5":
            job_searcher.searchByDate()
        case "bk":
            mainMenu(worksheet)


def updateJobDetail():
    checkOs()
    pass


# [DONE] GENERATE TITLE AND ALL COLUMN NAMES
def generateTitles(worksheet):
    heading_fill = PatternFill(fill_type='solid', start_color='87f5a4', end_color='87f5a4')
    heading_font = Font(size=16, bold=True)
    for i, n in zip(upperCol, colName):
        cell = i + "2"
        worksheet[cell] = n

    # after filling row 1 with your headers
    for col_idx, header in enumerate(colName, start=1):
        column_letter = get_column_letter(col_idx)
        # set width a bit bigger than the header length
        worksheet.column_dimensions[column_letter].width = len(header) + 2

    worksheet.merge_cells('A1:I1')
    worksheet['A1'] = "Job Tracker"
    worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    worksheet['A1'].font = heading_font
    worksheet['A1'].fill = heading_fill


def mainMenu(worksheet):
    while True:
        print("\n|  JOB TRACKER MENU  | \n\n"
              "1. Enter Job Entry \n" +
              "2. Search Job Entry \n" +
              "3. Update Job Entry \n" +
              "4. Quit Program")
        opts = input("> ")
        print()

        match opts:
            case "1":
                getJobDetails(worksheet)
            case "2":
                searchJobDetail(worksheet)
            case "3":
                updateJobDetail()
            case "4":
                print("Exiting program.")
                sys.exit()
                break
            case _:
                print("Invalid Input")


# [ADDED] CREATED AN ENTRY POINT FOR ALL CALLS AND CODE TO BE IN
# SINCE ITS NOT CLASS BASED, IM NOT CREATING A MAIN FUNCTION
if __name__ == "__main__":
    try:
        wb = checkWorksheet()
        ws = wb.active
        mainMenu(ws)
    except KeyboardInterrupt:
        print("\n")
    finally:
        # Save the file
        wb.save("Job-Tracker.xlsx")
        print("workload saved")
